from __future__ import annotations

import shutil
import tempfile
import zipfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree

from .xlsx_csv import CELL_REF_RE, NS, XlsxCsvError, choose_sheet, read_sheets, read_shared_strings, read_text


class XlsxEditError(Exception):
    """Ошибка анализа или редактирования XLSX."""


@dataclass(frozen=True)
class ColumnInfo:
    # letter - Excel-адрес колонки, например A или AA. header - значение первой
    # строки. filled_cells оставлен для совместимости, но больше не считается
    # по всей колонке: Check columns теперь читает только строку заголовков.
    letter: str
    index: int
    header: str
    filled_cells: int


@dataclass(frozen=True)
class WorkbookColumnReport:
    source: Path
    sheet_name: str
    filled_column_count: int
    columns: list[ColumnInfo]


@dataclass(frozen=True)
class ColumnInsertResult:
    source: Path
    column: str
    sheet_name: str


@dataclass(frozen=True)
class ColumnDeleteResult:
    source: Path
    column: str
    sheet_name: str


@dataclass(frozen=True)
class ColumnSwapResult:
    source: Path
    first_column: str
    second_column: str
    sheet_name: str


def inspect_xlsx_columns(source: Path, *, sheet: str | int | None = None) -> WorkbookColumnReport:
    """Вернуть список непустых колонок выбранного листа."""

    source = Path(source)
    if not source.exists():
        raise XlsxEditError(f"File does not exist: {source}")

    try:
        with zipfile.ZipFile(source) as archive:
            shared_strings = read_shared_strings(archive)
            sheet_info = choose_sheet(read_sheets(archive), sheet)
            root = ElementTree.fromstring(archive.read(sheet_info.path))
            columns = collect_header_columns(root, shared_strings)
        return WorkbookColumnReport(
            source=source,
            sheet_name=sheet_info.name,
            filled_column_count=len(columns),
            columns=columns,
        )
    except (zipfile.BadZipFile, KeyError, ElementTree.ParseError, XlsxCsvError) as exc:
        raise XlsxEditError(f"Cannot inspect XLSX columns: {exc}") from exc


def inspect_xlsx_path_columns(
    source: Path,
    *,
    sheet: str | int | None = None,
    recursive: bool = False,
) -> list[WorkbookColumnReport]:
    """Проверить колонки в одном XLSX или во всех XLSX файлах папки."""

    files = resolve_xlsx_files(source, recursive=recursive)
    return [inspect_xlsx_columns(path, sheet=sheet) for path in files]


def insert_column_in_xlsx(
    source: Path,
    column: str,
    *,
    sheet: str | int | None = None,
    header: str = "",
) -> ColumnInsertResult:
    """Вставить пустую колонку в один XLSX файл на месте column."""

    source = Path(source)
    column_index = column_name_to_index(column) + 1
    try:
        load_workbook = get_openpyxl_load_workbook()
        workbook = load_workbook(source)
        worksheet = choose_worksheet(workbook, sheet)
        worksheet.insert_cols(column_index)
        if header:
            worksheet.cell(row=1, column=column_index).value = header
        save_workbook_atomically(workbook, source)
        return ColumnInsertResult(source=source, column=index_to_column_name(column_index - 1), sheet_name=worksheet.title)
    except Exception as exc:
        raise XlsxEditError(f"Cannot insert XLSX column: {exc}") from exc
    finally:
        close_workbook_if_possible(locals().get("workbook"))


def insert_column_in_xlsx_path(
    source: Path,
    column: str,
    *,
    sheet: str | int | None = None,
    header: str = "",
    recursive: bool = False,
) -> list[ColumnInsertResult]:
    """Вставить колонку в один файл или пачку XLSX файлов."""

    files = resolve_xlsx_files(source, recursive=recursive)
    return [insert_column_in_xlsx(path, column, sheet=sheet, header=header) for path in files]


def delete_column_in_xlsx(
    source: Path,
    column: str,
    *,
    sheet: str | int | None = None,
) -> ColumnDeleteResult:
    """Удалить колонку из одного XLSX файла."""

    source = Path(source)
    column_index = column_name_to_index(column) + 1
    try:
        load_workbook = get_openpyxl_load_workbook()
        workbook = load_workbook(source)
        worksheet = choose_worksheet(workbook, sheet)
        worksheet.delete_cols(column_index)
        save_workbook_atomically(workbook, source)
        return ColumnDeleteResult(source=source, column=index_to_column_name(column_index - 1), sheet_name=worksheet.title)
    except Exception as exc:
        raise XlsxEditError(f"Cannot delete XLSX column: {exc}") from exc
    finally:
        close_workbook_if_possible(locals().get("workbook"))


def delete_column_in_xlsx_path(
    source: Path,
    column: str,
    *,
    sheet: str | int | None = None,
    recursive: bool = False,
) -> list[ColumnDeleteResult]:
    """Удалить колонку в одном файле или пачке XLSX файлов."""

    files = resolve_xlsx_files(source, recursive=recursive)
    return [delete_column_in_xlsx(path, column, sheet=sheet) for path in files]


def swap_columns_in_xlsx(
    source: Path,
    first_column: str,
    second_column: str,
    *,
    sheet: str | int | None = None,
) -> ColumnSwapResult:
    """Поменять местами две колонки в одном XLSX файле."""

    source = Path(source)
    first_index = column_name_to_index(first_column) + 1
    second_index = column_name_to_index(second_column) + 1
    if first_index == second_index:
        raise XlsxEditError("Columns to swap must be different.")

    try:
        load_workbook = get_openpyxl_load_workbook()
        workbook = load_workbook(source)
        worksheet = choose_worksheet(workbook, sheet)
        swap_worksheet_columns(worksheet, first_index, second_index)
        save_workbook_atomically(workbook, source)
        return ColumnSwapResult(
            source=source,
            first_column=index_to_column_name(first_index - 1),
            second_column=index_to_column_name(second_index - 1),
            sheet_name=worksheet.title,
        )
    except Exception as exc:
        raise XlsxEditError(f"Cannot swap XLSX columns: {exc}") from exc
    finally:
        close_workbook_if_possible(locals().get("workbook"))


def swap_columns_in_xlsx_path(
    source: Path,
    first_column: str,
    second_column: str,
    *,
    sheet: str | int | None = None,
    recursive: bool = False,
) -> list[ColumnSwapResult]:
    """Поменять местами две колонки в одном файле или пачке XLSX файлов."""

    files = resolve_xlsx_files(source, recursive=recursive)
    return [swap_columns_in_xlsx(path, first_column, second_column, sheet=sheet) for path in files]


def swap_worksheet_columns(worksheet, first_index: int, second_index: int) -> None:
    max_row = max(worksheet.max_row, 1)
    for row in range(1, max_row + 1):
        first_cell = worksheet.cell(row=row, column=first_index)
        second_cell = worksheet.cell(row=row, column=second_index)
        swap_cells(first_cell, second_cell)

    first_letter = index_to_column_name(first_index - 1)
    second_letter = index_to_column_name(second_index - 1)
    first_dimension = copy(worksheet.column_dimensions[first_letter])
    second_dimension = copy(worksheet.column_dimensions[second_letter])
    worksheet.column_dimensions[first_letter] = second_dimension
    worksheet.column_dimensions[second_letter] = first_dimension


def swap_cells(first_cell, second_cell) -> None:
    first_snapshot = snapshot_cell(first_cell)
    second_snapshot = snapshot_cell(second_cell)
    apply_cell_snapshot(first_cell, second_snapshot)
    apply_cell_snapshot(second_cell, first_snapshot)


def snapshot_cell(cell) -> dict[str, object]:
    return {
        "value": cell.value,
        "style": copy(cell._style),
        "number_format": cell.number_format,
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "comment": copy(cell.comment) if cell.comment is not None else None,
        "hyperlink": copy(cell.hyperlink) if cell.hyperlink is not None else None,
    }


def apply_cell_snapshot(cell, snapshot: dict[str, object]) -> None:
    cell.value = snapshot["value"]
    cell._style = snapshot["style"]
    cell.number_format = snapshot["number_format"]
    cell.font = snapshot["font"]
    cell.fill = snapshot["fill"]
    cell.border = snapshot["border"]
    cell.alignment = snapshot["alignment"]
    cell.protection = snapshot["protection"]
    cell.comment = snapshot["comment"]
    cell._hyperlink = snapshot["hyperlink"]


def collect_header_columns(root: ElementTree.Element, shared_strings: list[str]) -> list[ColumnInfo]:
    # Самый быстрый путь для Check columns: читаем только row r="1" из XML
    # листа и не строим объектную модель всей книги.
    columns: list[ColumnInfo] = []
    first_row = root.find("m:sheetData/m:row[@r='1']", NS)
    if first_row is None:
        first_row = root.find("m:sheetData/m:row", NS)
    if first_row is None:
        return columns

    for cell in first_row.findall("m:c", NS):
        header = read_header_cell(cell, shared_strings)
        if header == "":
            continue
        index = cell_column_index(cell)
        columns.append(
            ColumnInfo(
                letter=index_to_column_name(index),
                index=index,
                header=header,
                filled_cells=1,
            )
        )
    return columns


def read_header_cell(cell: ElementTree.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)
    value = value_node.text if value_node is not None and value_node.text is not None else ""
    if cell_type == "s" and value:
        return shared_strings[int(value)]
    if cell_type == "inlineStr":
        return read_text(cell.find("m:is", NS))
    return value


def cell_column_index(cell: ElementTree.Element) -> int:
    reference = cell.attrib.get("r")
    if not reference:
        raise XlsxEditError("Header cell has no reference.")
    match = CELL_REF_RE.fullmatch(reference)
    if not match:
        raise XlsxEditError(f"Invalid cell reference: {reference}")
    return column_name_to_index(match.group("column"))


def choose_worksheet(workbook, sheet: str | int | None):
    # Пользователь может оставить лист пустым, указать имя или 1-based номер.
    if sheet is None or sheet == "":
        return workbook.active
    if isinstance(sheet, int) or str(sheet).isdigit():
        index = int(sheet) - 1
        if 0 <= index < len(workbook.worksheets):
            return workbook.worksheets[index]
        raise XlsxEditError(f"Sheet index is out of range: {sheet}")
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    raise XlsxEditError(f"Sheet not found: {sheet}")


def get_openpyxl_load_workbook():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise XlsxEditError(
            "XLSX write operations require openpyxl. Install it with: python -m pip install -r requirements.txt"
        ) from exc
    return load_workbook


def resolve_xlsx_files(source: Path, *, recursive: bool) -> list[Path]:
    source = Path(source)
    if source.is_file():
        return [source]
    if source.is_dir():
        files = list(iter_xlsx_files_for_edit(source, recursive=recursive))
        if files:
            return files
        raise XlsxEditError(f"No .xlsx files found in: {source}")
    raise XlsxEditError(f"Source must be an XLSX file or directory: {source}")


def iter_xlsx_files_for_edit(directory: Path, *, recursive: bool) -> Iterable[Path]:
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    for path in sorted(directory.glob(pattern)):
        if path.is_file() and not path.name.startswith("~$"):
            yield path


def save_workbook_atomically(workbook, source: Path) -> None:
    # Сначала сохраняем во временный файл рядом с оригиналом. Если сохранение
    # упадет, исходный XLSX останется на месте; после успеха заменяем файл.
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=source.parent) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        workbook.save(temp_path)
        shutil.move(str(temp_path), source)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def close_workbook_if_possible(workbook) -> None:
    if workbook is not None and hasattr(workbook, "close"):
        workbook.close()


def column_name_to_index(column: str) -> int:
    cleaned = column.strip().upper()
    if not cleaned or not cleaned.isalpha():
        raise XlsxEditError(f"Invalid column name: {column}")
    index = 0
    for char in cleaned:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def index_to_column_name(index: int) -> str:
    if index < 0:
        raise XlsxEditError(f"Invalid column index: {index}")
    result = ""
    value = index + 1
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result
